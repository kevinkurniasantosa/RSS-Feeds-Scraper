# For scraping
import requests
from bs4 import BeautifulSoup
import lxml
import urllib
import json
# For data manipulation
import pandas as pd 
import numpy as np 
# For read and write to MongoDB
import pymongo
# For password input
import stdiomask
# For using excel
import openpyxl
# For scraping news content with no RSS feed
from news_scraper_lib import *
import time

print('import success')

####################################################
#################################################### DEFINE GLOBAL VARIABLE

### THIS IS CUSTOMIZABLE

## NEWS RSS FEED
# list_rss_feed = [
#     'http://rss.cnn.com/rss/cnn_health.rss',
#     'http://feeds.bbci.co.uk/news/health/rss.xml',
#     'http://rssfeeds.webmd.com/rss/rss.aspx?RSSSource=RSS_PUBLIC'
# ]
list_rss_feed = []

## NEWS SOURCE NAME
# list_source_name = [
#     'cnn_health',
#     'bbc_health',
#     'webmd'
# ]
list_source_name = []

## EXCEL FILE CREDENTIALS
excel_filename = 'rss_criteria.xlsx'
excel_sheet_name = 'RSS Criteria'
starting_column = '3'
row_rss_feed_url = 'B'
row_rss_source_name = 'C'

## MONGODB CREDENTIAlS
username = 'kevin.santosa@ams-global.ai'
# password = ''
db_name = 'news'

#################################################### DEFINE FUNCTIONS

# For cleaning string value
def clean_string(x):
    x = x.replace('\"', '\'\'').replace('\r', ' ').replace('\n', ' ')
    x = unicodedata.normalize('NFKD', x).encode('ascii', 'ignore')
    x = x.decode('ascii')
    
    return x

# For collecting criteras in the excel file
def read_excel():
    # Connect to excel
    wb = openpyxl.load_workbook(excel_filename)
    sheet = wb["{}".format(excel_sheet_name)]

    print('Column len: ' + str(sheet.max_row))

    for i in range(int(starting_column), sheet.max_row+1):
        rss_feed_url = sheet[row_rss_feed_url+str(i)].value
        rss_source_name = sheet[row_rss_source_name+str(i)].value

        list_rss_feed.append(rss_feed_url)
        list_source_name.append(rss_source_name)
    
    print('List RSS feed: ' + str(list_rss_feed))
    print('List source name: ' + str(list_source_name))
    print('===============================================================')

    wb.close()

# For scraping RSS feed
def scraping_rss_feed(rss_feed):
    try:
        article_list_per_rss = []
        res = requests.get(rss_feed, headers={'User-Agent': 'Mozilla/5.0'})
        print('Request status: ' + str(res.status_code))
        soup = BeautifulSoup(res.content, 'xml')
        # print(soup.prettify())

        articles = soup.find_all('item')
        # print(articles)
        print('There are {} articles in {}'.format(len(articles), rss_feed))
        print('-------------------------------------')

        # Loop through the items/articles
        for article in articles:
            # Get link
            try:
                link = article.find('link').text
            except:
                link = '-'
            print('Link: ' + str(link))

            # Get title
            try:
                title = article.find('title').text.strip()
            except:
                title = '-'
            print('Title: ' + str(title))

            # Get published date
            try:
                published_date = article.find('pubDate').text.strip()
            except:
                published_date = '-'
            print('Published date: ' + str(published_date))

            # Get description
            try:
                description = article.find('description').text.strip()
            except:
                description = '-'
            print('Description: ' + str(description))

            # Get content (news content is gotten from scraping into the news sources - not through RSS feed since 
            # public RSS feed does not provide the content of the news) 
            # note: if there is no link, there will be no content

            # try:
            #     content = ''
            #     article_url = link
            #     print(article_url)
            #     res = requests.get(article_url, headers={'User-Agent': 'Mozilla/5.0'})
            #     soup_article = BeautifulSoup(res.text, 'html.parser')
            #     if 'bbc.co.uk' in article_url:
            #         content = scraping_bbc_content(soup_article)
            #     elif 'liputan6.com' in article_url:
            #         content = scraping_liputan6_content(soup_article)
            # except:
            #     content = '-'
            # print('Content: ' + str(content))

            # Store in a dictionary
            article_dict = {
                'link': link,
                'title': title,
                'published_date': published_date,
                'description': description
            }
            
            article_list_per_rss.append(article_dict)
            print('-------------------------')

    except Exception as err:
        print('Scraping failed, see exception as follows: ' + str(err))

    return article_list_per_rss

## For scraping RSS feed
def read_and_write_rss_to_json():
    # Loop through RSS feed list
    for i, rss_feed in enumerate(list_rss_feed):
        print('========================================= {}'.format(list_source_name[i].upper()))

        source_name = list_source_name[i].upper()
        source_name = source_name.replace('_', ' ')
        print('Scraping {}'.format(source_name))

        print('-------------------------------------')
        articles_per_rss = scraping_rss_feed(rss_feed)
        # print(articles_per_rss)
        
        # Write to JSON data
        list_source_name[i] = list_source_name[i].replace(' ', '_') # clean string
        with open('{}_rss.json'.format(list_source_name[i].lower()), 'w') as outfile:
            json.dump(articles_per_rss, outfile, indent=2)
            print('File {}_articles.txt successfuly created!'.format(list_source_name[i]))

            
## For scraping BBC news content
def scraping_bbc_content(soup_article):
    article_content = soup_article.find('article')
    # print(article_content.prettify())
    list_sentences = article_content.find_all('div', {'data-component': 'text-block'})
    content = list_sentences[0] + list_sentences[1] + list_sentences[2]

    return content

## For scraping Liputan6 news content
def scraping_liputan6_content(soup_article):
    article_content = soup_article.find('article')
    content = article_content.find('div', class_='read-page--content')
    # print(article_content.prettify())

    return content

## Store data to MongoDB
def store_to_db():
    password = stdiomask.getpass()

    # Connect to MongoDB
    client = pymongo.MongoClient("mongodb+srv://{0}:{1}@cluster0.j2t0m.mongodb.net/{2}?retryWrites=true&w=majority".format(urllib.parse.quote(username), password, db_name))
    client = pymongo.MongoClient('112.215.45.140/32', 27017, username="my_name", password='my_pwd')
    # Specify particular database and collection
    db = client.news
    collection = db.news_scraping
    
    # Sample document
    document1 = {
        "link": "http://rss.cnn.com/~r/rss/cnn_health/~3/G7tBayBZaYo/index.html",
        "title": "Most Americans probably won't be able to get a Covid-19 vaccine until mid-2021, CDC director says",
        "published_date": "Thu, 17 Sep 2020 08:43:27 GMT",
        "description": "A wedding in Maine is linked to 176 Covid-19 cases and the deaths of seven people who didn't attend the celebration, demonstrating  just how easily and quickly the virus can spread at social gatherings, public health experts say.<div class=\"feedflare\">\n<a href=\"http://rss.cnn.com/~ff/rss/cnn_health?a=G7tBayBZaYo:1hj3ZE5cRdQ:yIl2AUoC8zA\"><img src=\"http://feeds.feedburner.com/~ff/rss/cnn_health?d=yIl2AUoC8zA\" border=\"0\"></img></a> <a href=\"http://rss.cnn.com/~ff/rss/cnn_health?a=G7tBayBZaYo:1hj3ZE5cRdQ:7Q72WNTAKBA\"><img src=\"http://feeds.feedburner.com/~ff/rss/cnn_health?d=7Q72WNTAKBA\" border=\"0\"></img></a> <a href=\"http://rss.cnn.com/~ff/rss/cnn_health?a=G7tBayBZaYo:1hj3ZE5cRdQ:V_sGLiPBpWU\"><img src=\"http://feeds.feedburner.com/~ff/rss/cnn_health?i=G7tBayBZaYo:1hj3ZE5cRdQ:V_sGLiPBpWU\" border=\"0\"></img></a> <a href=\"http://rss.cnn.com/~ff/rss/cnn_health?a=G7tBayBZaYo:1hj3ZE5cRdQ:qj6IDK7rITs\"><img src=\"http://feeds.feedburner.com/~ff/rss/cnn_health?d=qj6IDK7rITs\" border=\"0\"></img></a> <a href=\"http://rss.cnn.com/~ff/rss/cnn_health?a=G7tBayBZaYo:1hj3ZE5cRdQ:gIN9vFwOqvQ\"><img src=\"http://feeds.feedburner.com/~ff/rss/cnn_health?i=G7tBayBZaYo:1hj3ZE5cRdQ:gIN9vFwOqvQ\" border=\"0\"></img></a>\n</div><img src=\"http://feeds.feedburner.com/~r/rss/cnn_health/~4/G7tBayBZaYo\" height=\"1\" width=\"1\" alt=\"\"/>"             
    }
    
    print(collection)
    # collection.insert_one(document1)
    # collection.insert_many()

    # Printing the data inserted
    cursor = collection.find()
    for record in cursor:
        print(record)
    
## Main function
def main():
    # Read excel
    read_excel()

    # Collect RSS feed
    read_and_write_rss_to_json()

    # Store RSS feed data to database
    # store_to_db()

#################################################### START HERE

if __name__ == '__main__':
    main()